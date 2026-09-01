import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def text(value) -> str:
    return "" if value is None else str(value).strip()


def normalize_header(value) -> str:
    return re.sub(r"\s+", " ", text(value)).casefold()


def jcp_prefix(value: str) -> str:
    raw = text(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 7:
        return digits[:7]
    return raw[:7]


def find_column(headers: list[str], candidates: list[str], fallback_index: int) -> int:
    normalized = [normalize_header(header) for header in headers]
    for candidate in candidates:
        key = normalize_header(candidate)
        for index, header in enumerate(normalized):
            if header == key:
                return index
    for candidate in candidates:
        key = normalize_header(candidate)
        for index, header in enumerate(normalized):
            if key and key in header:
                return index
    return fallback_index


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8", newline="\n")
    parser = argparse.ArgumentParser(description="Prepare JCP source rows and SKU name + prefix groups from an Excel workbook.")
    parser.add_argument("--source-xlsx", required=True, help="Input workbook containing source SKU, seller SKU, JCP SKU, and SKU/SPU name.")
    parser.add_argument("--output-dir", required=True, help="Directory where source_rows.json should be written.")
    parser.add_argument("--output-json", default="", help="Optional explicit source_rows.json output path.")
    args = parser.parse_args()

    source = Path(args.source_xlsx).resolve()
    output_dir = Path(args.output_dir).resolve()
    output = Path(args.output_json).resolve() if args.output_json else output_dir / "source_rows.json"

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [text(cell.value) for cell in header_row]

    sku_index = find_column(headers, ["SKU"], 0)
    seller_sku_index = find_column(headers, ["Seller Sku", "Seller SKU"], 1)
    crawl_sku_index = find_column(headers, ["SKU（用于爬取）", "JCP SKU", "Crawl SKU"], 2)
    sku_name_index = find_column(headers, ["SPU Name", "SKU Name", "sku name"], 3)

    rows = []
    for source_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(values or [])
        if not values or all(text(value) == "" for value in values):
            continue

        source_sku = text(values[sku_index] if sku_index < len(values) else "")
        seller_sku = text(values[seller_sku_index] if seller_sku_index < len(values) else "")
        crawl_sku = text(values[crawl_sku_index] if crawl_sku_index < len(values) else "")
        sku_name = text(values[sku_name_index] if sku_name_index < len(values) else "")
        if not crawl_sku or not sku_name:
            continue

        prefix = jcp_prefix(crawl_sku)
        rows.append(
            {
                "source_row": source_row,
                "source_sku": source_sku,
                "seller_sku": seller_sku,
                "crawl_sku": crawl_sku,
                "sku_name": sku_name,
                "spu_name": sku_name,
                "jcp_prefix": prefix,
                "source_row_values": [text(value) for value in values],
            }
        )

    groups_by_key = {}
    groups = []
    for row in rows:
        group_key = f"{row['sku_name']}|{row['jcp_prefix']}"
        if group_key not in groups_by_key:
            group = {
                "group_key": group_key,
                "sku_name": row["sku_name"],
                "spu_name": row["spu_name"],
                "jcp_prefix": row["jcp_prefix"],
                "sku_prefix": row["jcp_prefix"],
                "request_sku": row["crawl_sku"],
                "source_count": 0,
                "source_skus": [],
                "source_rows": [],
            }
            groups_by_key[group_key] = group
            groups.append(group)
        group = groups_by_key[group_key]
        group["source_count"] += 1
        group["source_skus"].append(row["crawl_sku"])
        group["source_rows"].append(row["source_row"])

    payload = {
        "source_workbook": str(source).replace("\\", "/"),
        "sheet": ws.title,
        "headers": headers,
        "column_mapping": {
            "source_sku": headers[sku_index] if sku_index < len(headers) else "",
            "seller_sku": headers[seller_sku_index] if seller_sku_index < len(headers) else "",
            "jcp_sku": headers[crawl_sku_index] if crawl_sku_index < len(headers) else "",
            "sku_name": headers[sku_name_index] if sku_name_index < len(headers) else "",
        },
        "grouping_rule": "First group by SKU/SPU name, then by the first 7 digits of the source JCP SKU / webId prefix. Each unique source JCP SKU is resolved through product-aliases, while summaries retain the SKU name + prefix business group.",
        "matching_rule": "After each deduped PDP is scraped, source rows are matched only by exact source JCP SKU == PDP variant sku_id within the same SKU name + prefix group. selectedSKUId is not used as a replacement match key, and missing rows are preserved.",
        "row_count": len(rows),
        "group_count": len(groups),
        "rows": rows,
        "groups": groups,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "source_workbook": payload["source_workbook"],
                "sheet": payload["sheet"],
                "headers": payload["headers"],
                "row_count": payload["row_count"],
                "group_count": payload["group_count"],
                "grouping_rule": payload["grouping_rule"],
                "output": str(output),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    for index, group in enumerate(groups, start=1):
        print(
            f"{index:02d}. {group['sku_name']} | {group['jcp_prefix']} -> "
            f"{group['request_sku']} ({group['source_count']} SKUs)"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
